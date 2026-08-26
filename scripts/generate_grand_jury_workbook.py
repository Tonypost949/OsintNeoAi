"""Generates a multi-sheet Court-Ready Excel Workbook for Grand Jury & DOJ submission."""

import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def build_grand_jury_workbook():
    wb = openpyxl.Workbook()
    
    # -------------------------------------------------------------
    # Styling Definitions
    # -------------------------------------------------------------
    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    accent_fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
    gold_fill = PatternFill(start_color="D97706", end_color="D97706", fill_type="solid")
    green_fill = PatternFill(start_color="059669", end_color="059669", fill_type="solid")
    
    title_font = Font(name="Arial", size=14, bold=True, color="FFFFFF")
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    bold_font = Font(name="Arial", size=10, bold=True, color="000000")
    regular_font = Font(name="Arial", size=10, color="1E293B")
    
    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )

    # -------------------------------------------------------------
    # Sheet 1: Executive Summary
    # -------------------------------------------------------------
    ws1 = wb.active
    ws1.title = "Executive Summary & Parties"
    ws1.views.sheetView[0].showGridLines = True
    
    ws1.merge_cells("A1:F1")
    ws1["A1"] = "FORMAL EVIDENTIARY SUBMISSION TO THE ORANGE COUNTY GRAND JURY & U.S. DEPARTMENT OF JUSTICE"
    ws1["A1"].font = title_font
    ws1["A1"].fill = header_fill
    ws1["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[1].height = 40

    metadata = [
        ("SUBMISSION TYPE", "Qui Tam False Claims Act (31 U.S.C. § 3730) & Criminal Referral (18 U.S.C. § 1346 / § 1962)"),
        ("RELATOR / CLAIMANT", "Anthony Michael DiMarcello III (u/OSINTNeoAi / amd949609@gmail.com)"),
        ("PRIMARY TARGETS", "Shea Properties / Arte Moreno (SRB Management) / Former Mayor Harry Sidhu / Todd Ament / City of Anaheim"),
        ("PRIMARY STATUTES", "Cal. Surplus Land Act (Gov. Code § 54220), Cal. Civ. Code § 1942.5, Cal. Labor Code § 1102.5, 18 U.S.C. § 1346"),
        ("PRIMARY TIME HORIZON", "January 2019 – Present (Critical 24-Hour Collapse: May 23–24, 2022)"),
        ("TOTAL RECOVERY VALUATION", "$96,400,000 to $196,300,000+ (Statutory Bounties, Treble Damages & Punitive Multipliers)")
    ]

    for idx, (k, v) in enumerate(metadata, start=3):
        ws1[f"A{idx}"] = k
        ws1[f"A{idx}"].font = bold_font
        ws1[f"A{idx}"].fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
        ws1[f"A{idx}"].border = thin_border
        
        ws1.merge_cells(f"B{idx}:F{idx}")
        ws1[f"B{idx}"] = v
        ws1[f"B{idx}"].font = regular_font
        ws1[f"B{idx}"].border = thin_border
        ws1.row_dimensions[idx].height = 24

    # -------------------------------------------------------------
    # Sheet 2: Master Exhibit Ledger
    # -------------------------------------------------------------
    ws2 = wb.create_sheet(title="Master Exhibit Ledger")
    ws2.views.sheetView[0].showGridLines = True
    
    headers2 = ["Exhibit Code", "Forensic Topic", "Document Title & Description", "Governing Statute", "Verification Status", "Primary Source Citation"]
    ws2.append(headers2)
    ws2.row_dimensions[1].height = 28
    for col_idx in range(1, len(headers2) + 1):
        cell = ws2.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    exhibits_data = [
        ("EXHIBIT A-1", "Notice Defect", "15-Day Notice acknowledging COVID financial distress receipt (March 18, 2021)", "SB 91 / AB 832 Safe Harbor", "VERIFIED ADMISSION", "EV-021 / Landlord File"),
        ("EXHIBIT A-2", "Perjured Verification", "Shea Assistant Community Manager false verification Box 7d(2) (May 15, 2021)", "Cal. Penal Code § 118", "VERIFIED PERJURY", "UD-101 Verification (Case 30-2021-01201327)"),
        ("EXHIBIT A-3", "Armed Dispossession", "OCSD Levying Officer File No. 2021102780 (8 armed deputies lockout on Aug 4, 2021)", "Cal. Civ. Code § 1942.5", "VERIFIED SHERIFF RECORD", "Levying Officer Notice / Barnes Command"),
        ("EXHIBIT B-1", "Shelter Toxic Cap", "OCHCA Case Closed Certificate 20IC002 issued without DTSC concurrence (Aug 21, 2020)", "RCRA / CERCLA § 107", "VERIFIED REGULATORY FRAUD", "OCHCA Certificate 20IC002"),
        ("EXHIBIT B-2", "Toxic Assay", "Soil assays showing Hexavalent Chromium at 490 ppb (49x EPA limit) at HBNC shelter", "Prop 65 / Cal. H&S § 25249.5", "VERIFIED LAB ASSAY", "TOX-001 / TOX-002 Assays"),
        ("EXHIBIT C-1", "HCD Notice of Violation", "State Housing Department $96,000,000 mandatory statutory fine demand", "AB 1486 / Gov. Code § 54230.5", "VERIFIED NOTICE", "HCD Dec 8, 2021 Notice of Violation"),
        ("EXHIBIT C-2", "FBI Wiretap Affidavits", "Special Agent Brian Adkins affidavit detailing Sidhu $1M bribe & appraisal leaks", "18 U.S.C. § 1346 / § 1001", "VERIFIED PLEA RECORD", "USA v. Sidhu (Case 8:23-cr-00115-DOC)"),
        ("EXHIBIT D-1", "Whistleblower Transmission", "Whistleblower transmission 'Shea, The Angels, Moreno and Roundtree' (May 23, 2022)", "Cal. Labor Code § 1102.5", "VERIFIED RFC822 IMAP", "Gmail IMAP Header MID 19e928d913ea9a22"),
        ("EXHIBIT D-2", "Unanimous Kill Vote", "Anaheim City Council 7-0 vote terminating $320M stadium purchase agreement", "Honest Services Rescission", "VERIFIED COUNCIL MINUTES", "Anaheim City Council Official Minutes (May 24, 2022)")
    ]

    for row_idx, row in enumerate(exhibits_data, start=2):
        ws2.append(row)
        ws2.row_dimensions[row_idx].height = 24
        for col_idx in range(1, len(row) + 1):
            cell = ws2.cell(row=row_idx, column=col_idx)
            cell.font = regular_font
            cell.border = thin_border
            if col_idx == 5:
                cell.font = Font(name="Arial", size=9, bold=True, color="059669")
                cell.alignment = Alignment(horizontal="center")

    # -------------------------------------------------------------
    # Sheet 3: Damages Valuation Model
    # -------------------------------------------------------------
    ws3 = wb.create_sheet(title="Damages Valuation Matrix")
    ws3.views.sheetView[0].showGridLines = True
    
    headers3 = ["Pillar Code", "Recovery Category", "Governing Legal Framework", "Forensic Basis", "Low-End Recovery", "Ceiling Exposure"]
    ws3.append(headers3)
    ws3.row_dimensions[1].height = 28
    for col_idx in range(1, len(headers3) + 1):
        cell = ws3.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    damages_data = [
        ("PILLAR 1", "Qui Tam Relator Bounty (HCD Fine)", "31 U.S.C. § 3730(d) & Cal. FCA § 12652", "15%-30% of $96M statutory penalty levied by HCD", "$14,400,000", "$28,800,000"),
        ("PILLAR 2", "Qui Tam Relator Bounty (Land Valuation)", "31 U.S.C. § 3730(d) & Cal. FCA § 12652", "15%-30% of $320M fraudulent land transaction valuation", "$48,000,000", "$96,000,000"),
        ("PILLAR 3", "Tenant Retaliation & Eviction Fraud", "Cal. Civ. Code § 1942.5 & Emergency Acts", "Treble damages, statutory penalties & business loss", "$1,500,000", "$3,500,000"),
        ("PILLAR 4", "Civil Rights & Medical Endangerment", "Unruh Act (Civ. Code § 51) / IIED", "Displacement of disabled dependent mother from Hoag care", "$7,500,000", "$18,000,000"),
        ("PILLAR 5", "Punitive & Exemplary Damages", "Cal. Civ. Code § 3294 (Oppression & Malice)", "Premeditated corporate conspiracy & witness suppression", "$25,000,000", "$50,000,000"),
        ("TOTAL", "AGGREGATED RELATOR RECOVERY CEILING", "Comprehensive Multi-Pillar Valuation", "Full Statutory, Compensatory & Exemplary Total", "$96,400,000", "$196,300,000+")
    ]

    for row_idx, row in enumerate(damages_data, start=2):
        ws3.append(row)
        ws3.row_dimensions[row_idx].height = 24
        for col_idx in range(1, len(row) + 1):
            cell = ws3.cell(row=row_idx, column=col_idx)
            cell.font = regular_font
            cell.border = thin_border
            if row[0] == "TOTAL":
                cell.font = Font(name="Arial", size=11, bold=True, color="1E293B")
                cell.fill = PatternFill(start_color="FEF08A", end_color="FEF08A", fill_type="solid")
            elif col_idx in [5, 6]:
                cell.font = Font(name="Arial", size=10, bold=True, color="1E293B")
                cell.alignment = Alignment(horizontal="right")

    # Auto-adjust column widths for all sheets
    for sheet in [ws1, ws2, ws3]:
        for col in sheet.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            sheet.column_dimensions[col_letter].width = max(max_len + 4, 15)

    root_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    out_dir = os.path.join(root_dir, "exports")
    os.makedirs(out_dir, exist_ok=True)
    out_file = os.path.join(out_dir, "GRAND_JURY_DOJ_EVIDENTIARY_PACKET_2026.xlsx")
    wb.save(out_file)
    print(f"[✓] Successfully generated Court-Ready Excel Workbook: {out_file}")
    return out_file

if __name__ == "__main__":
    build_grand_jury_workbook()
