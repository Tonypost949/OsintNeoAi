#!/usr/bin/env python3
"""
OSINT AI Neo — Android Phone Scanner v2 (Deep Extraction)
Run this in Termux on Android. Read-only — never modifies your files.
Extracts: emails, phones, names, addresses, GPS, EXIF, audio tags,
          case numbers, keywords from every readable file type.
"""

import os, sys, json, hashlib, csv, re, mimetypes
from datetime import datetime
from pathlib import Path

OUTPUT_DIR  = "/storage/emulated/0/OSINT_Neo"
MASTER_XLSX = "/storage/emulated/0/OSINT_Neo/OSINT_Master.xlsx"
MASTER_LOG  = "/storage/emulated/0/OSINT_Neo/master_index.json"
PHONE_ROOT  = "/storage/emulated/0"

try:
    from PIL import Image
    from PIL.ExifTags import TAGS, GPSTAGS
    PIL_OK = True
except ImportError:
    PIL_OK = False
    print("[WARN] Pillow not installed — pip install Pillow")

try:
    import mutagen
    MUTAGEN_OK = True
except ImportError:
    MUTAGEN_OK = False
    print("[WARN] Mutagen not installed — pip install mutagen")

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False
    print("[WARN] openpyxl not installed — pip install openpyxl")

# ── File Categories ───────────────────────────────────────────────────────────
CATEGORIES = {
    "image":    [".jpg", ".jpeg", ".png", ".gif", ".bmp", ".tiff", ".webp", ".heic", ".raw"],
    "video":    [".mp4", ".mov", ".avi", ".mkv", ".wmv", ".3gp", ".m4v", ".flv"],
    "audio":    [".mp3", ".m4a", ".wav", ".flac", ".aac", ".ogg", ".wma", ".opus"],
    "document": [".pdf", ".doc", ".docx", ".txt", ".csv", ".xls", ".xlsx", ".odt", ".rtf", ".pptx", ".html", ".htm"],
    "contact":  [".vcf", ".vcard"],
    "data":     [".json", ".xml", ".yaml", ".yml", ".db", ".sqlite", ".sql"],
    "apk":      [".apk", ".xapk"],
    "archive":  [".zip", ".rar", ".7z", ".tar", ".gz"],
}

RISK_KEYWORDS = [
    "framed", "evidence", "criminal", "offense", "fraud", "bankruptcy",
    "lien", "court", "trial", "eviction", "whistleblower", "corruption",
    "qui tam", "patriot act", "sar", "displacement", "expose", "investigation",
    "ex parte", "legal argument", "defendant", "felony", "misdemeanor",
    "identity theft", "forgery", "perjury", "extortion", "bribery",
    "money laundering", "trafficking", "assault", "restraining order",
    "warrant", "subpoena", "deposition", "lawsuit", "plaintiff",
]

def cat(ext):
    e = ext.lower()
    for c, exts in CATEGORIES.items():
        if e in exts:
            return c
    return "other"

# ── Intelligence Extraction Patterns ─────────────────────────────────────────
RE_EMAIL   = re.compile(r'\b[A-Za-z0-9._%+\-]+@[A-Za-z0-9.\-]+\.[A-Za-z]{2,}\b')
RE_PHONE   = re.compile(r'\b(?:\+?1[\s.\-]?)?(?:\(?\d{3}\)?[\s.\-]?)?\d{3}[\s.\-]?\d{4}\b')
RE_SSN     = re.compile(r'\b\d{3}[-\s]?\d{2}[-\s]?\d{4}\b')
RE_CASE    = re.compile(r'\b(?:case\s*#?\s*|cacdce-|cv-|cr-|sc-|bc-)\w[\w\-]{3,}\b', re.I)
RE_DATE    = re.compile(r'\b(?:\d{1,2}[/\-]\d{1,2}[/\-]\d{2,4}|\d{4}[/\-]\d{2}[/\-]\d{2}|(?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)\w*\.?\s+\d{1,2},?\s+\d{4})\b', re.I)
RE_MONEY   = re.compile(r'\$[\d,]+(?:\.\d{2})?|\b\d[\d,]+\s*(?:dollars?|USD)\b', re.I)
RE_ADDRESS = re.compile(r'\d+\s+[A-Z][a-z]+(?:\s+[A-Z][a-z]+){0,3}\s+(?:St|Ave|Blvd|Dr|Rd|Ct|Ln|Way|Pl|Circle|Loop|Pkwy|Hwy)\.?(?:\s+(?:Apt|Suite|Unit|#)\s*\w+)?', re.I)
RE_NAME    = re.compile(r'\b([A-Z][a-z]{1,15})\s+(?:[A-Z]\.\s+)?([A-Z][a-z]{1,20})(?:\s+([A-Z][a-z]{1,20}))?\b')

STOP_WORDS = {
    "The", "This", "That", "With", "From", "Your", "Have", "Been", "Will",
    "More", "Some", "They", "When", "What", "Which", "Were", "Also", "Into",
    "About", "After", "Before", "Through", "During", "However", "Between",
    "Other", "Their", "There", "These", "Those", "Each", "Than", "Then",
    "Only", "Very", "Much", "Such", "Both", "Just", "Even", "Most",
    "File", "Page", "Date", "Time", "Type", "Name", "Case", "Code",
    "City", "State", "Court", "Legal", "Data", "Note", "Info", "List",
    "True", "False", "None", "Null",
}

def extract_names(text):
    found = []
    for m in RE_NAME.finditer(text):
        first, last = m.group(1), m.group(2)
        if first not in STOP_WORDS and last not in STOP_WORDS and len(first) > 1 and len(last) > 1:
            full = m.group(0).strip()
            if full not in found:
                found.append(full)
    return found[:10]

def extract_intelligence(text):
    """Run all regex extractors on a block of text."""
    intel = {}
    emails = list(set(RE_EMAIL.findall(text)))
    if emails:
        intel["emails_found"] = " | ".join(emails[:5])
    phones = list(set(RE_PHONE.findall(text)))
    phones = [p.strip() for p in phones if len(re.sub(r'\D','',p)) >= 10]
    if phones:
        intel["phones_found"] = " | ".join(phones[:5])
    ssns = list(set(RE_SSN.findall(text)))
    if ssns:
        intel["ssns_found"] = " | ".join(ssns[:3])
    cases = list(set(RE_CASE.findall(text)))
    if cases:
        intel["case_numbers"] = " | ".join(cases[:5])
    dates = list(set(RE_DATE.findall(text)))
    if dates:
        intel["dates_found"] = " | ".join(dates[:5])
    money = list(set(RE_MONEY.findall(text)))
    if money:
        intel["money_refs"] = " | ".join(money[:5])
    addresses = list(set(RE_ADDRESS.findall(text)))
    if addresses:
        intel["addresses_found"] = " | ".join(addresses[:3])
    names = extract_names(text)
    if names:
        intel["names_found"] = " | ".join(names)
    hits = [k for k in RISK_KEYWORDS if k in text.lower()]
    if hits:
        intel["keywords_hit"] = ", ".join(hits[:10])
        intel["risk_flag"] = "High" if len(hits) >= 2 else "Medium"
    return intel

# ── File Readers ──────────────────────────────────────────────────────────────
def read_text(path, max_chars=5000):
    try:
        with open(path, "r", encoding="utf-8", errors="ignore") as f:
            return f.read(max_chars)
    except Exception:
        return ""

def read_html(path):
    try:
        text = read_text(path, 8000)
        text = re.sub(r'<[^>]+>', ' ', text)
        text = re.sub(r'\s+', ' ', text)
        return text[:5000]
    except Exception:
        return ""

def md5(path):
    h = hashlib.md5()
    try:
        with open(path, "rb") as f:
            while chunk := f.read(8192):
                h.update(chunk)
        return h.hexdigest()
    except Exception:
        return ""

def gps_decimal(coord, ref):
    try:
        d, m, s = coord
        dec = float(d) + float(m)/60 + float(s)/3600
        if ref in ["S", "W"]:
            dec = -dec
        return round(dec, 6)
    except Exception:
        return None

def extract_image(path):
    meta = {}
    if not PIL_OK:
        return meta
    try:
        with Image.open(path) as img:
            meta["img_width"] = img.size[0]
            meta["img_height"] = img.size[1]
            exif = img._getexif()
            if exif:
                for tag_id, val in exif.items():
                    tag = TAGS.get(tag_id, tag_id)
                    if isinstance(val, bytes):
                        continue
                    if str(tag) == "GPSInfo":
                        gps = {GPSTAGS.get(k,k): v for k,v in val.items()}
                        lat = gps_decimal(gps.get("GPSLatitude",(0,0,0)), gps.get("GPSLatitudeRef","N"))
                        lon = gps_decimal(gps.get("GPSLongitude",(0,0,0)), gps.get("GPSLongitudeRef","E"))
                        if lat and lon:
                            meta["gps_lat"] = lat
                            meta["gps_lon"] = lon
                            meta["gps_location"] = f"https://maps.google.com/?q={lat},{lon}"
                    elif str(tag) in ["Make","Model","DateTime","DateTimeOriginal",
                                      "Software","Artist","Copyright","ImageDescription"]:
                        meta[f"exif_{tag}"] = str(val)[:200]
    except Exception as e:
        meta["img_error"] = str(e)
    return meta

def extract_audio(path):
    meta = {}
    if not MUTAGEN_OK:
        return meta
    try:
        audio = mutagen.File(path)
        if audio:
            tag_map = {
                "TIT2": "audio_title", "TPE1": "audio_artist", "TALB": "audio_album",
                "TDRC": "audio_year",  "TCON": "audio_genre",
                "title": "audio_title","artist": "audio_artist","album": "audio_album",
                "date": "audio_year",  "genre": "audio_genre",
            }
            for k, mapped in tag_map.items():
                if k in audio:
                    meta[mapped] = str(audio[k])[:200]
            if hasattr(audio, "info"):
                if hasattr(audio.info, "length"):
                    meta["audio_duration"] = f"{int(audio.info.length//60)}m{int(audio.info.length%60)}s"
                if hasattr(audio.info, "bitrate"):
                    meta["audio_bitrate"] = f"{audio.info.bitrate}kbps"
    except Exception as e:
        meta["audio_error"] = str(e)
    return meta

def extract_vcf(path):
    meta = {}
    try:
        content = read_text(path, 10000)
        phones, emails = [], []
        for line in content.splitlines():
            up = line.upper()
            if line.startswith("FN:"):
                meta["names_found"] = line[3:].strip()
            elif up.startswith("TEL"):
                phones.append(line.split(":")[-1].strip())
            elif up.startswith("EMAIL"):
                emails.append(line.split(":")[-1].strip())
            elif up.startswith("ADR"):
                meta["addresses_found"] = line.split(":")[-1].replace(";", " ").strip()
            elif line.startswith("ORG:"):
                meta["orgs_found"] = line[4:].strip()
            elif line.startswith("BDAY:"):
                meta["dates_found"] = line[5:].strip()
        if phones:
            meta["phones_found"] = " | ".join(phones[:5])
        if emails:
            meta["emails_found"] = " | ".join(emails[:5])
    except Exception as e:
        meta["vcf_error"] = str(e)
    return meta

# ── Main File Scanner ─────────────────────────────────────────────────────────
def scan_file(filepath):
    p = Path(filepath)
    if not p.is_file():
        return None
    ext   = p.suffix.lower()
    fcat  = cat(ext)
    try:
        stat     = p.stat()
        size     = stat.st_size
        modified = datetime.fromtimestamp(stat.st_mtime).strftime("%Y-%m-%d %H:%M:%S")
    except Exception:
        size, modified = 0, ""

    meta = {
        "file_name":  p.name,
        "file_path":  str(filepath),
        "folder":     str(p.parent),
        "file_type":  ext,
        "category":   fcat,
        "file_size":  size,
        "modified":   modified,
        "md5_hash":   md5(filepath),
        "scan_label": "Phone Scan",
        "risk_flag":  "Low",
    }

    # ── Type-specific extraction ──────────────────────────────────────────────
    if fcat == "image":
        meta.update(extract_image(filepath))

    elif fcat == "audio":
        meta.update(extract_audio(filepath))

    elif fcat == "contact":
        meta.update(extract_vcf(filepath))

    elif fcat == "document":
        if ext in [".txt", ".csv", ".rtf"]:
            text = read_text(filepath, 6000)
            meta["content_preview"] = text[:300].replace("\n", " ").strip()
            meta.update(extract_intelligence(text))

        elif ext in [".html", ".htm"]:
            text = read_html(filepath)
            meta["content_preview"] = text[:300].strip()
            meta.update(extract_intelligence(text))

        elif ext in [".json"]:
            text = read_text(filepath, 4000)
            meta.update(extract_intelligence(text))

        elif ext in [".pdf", ".docx", ".doc", ".odt"]:
            # Try to read as text (works for some PDFs/text-based docs)
            text = read_text(filepath, 6000)
            if len(text.strip()) > 50:
                meta["content_preview"] = text[:300].replace("\n", " ").strip()
                meta.update(extract_intelligence(text))
            # Also mine the filename itself
            fname_intel = extract_intelligence(p.stem.replace("-", " ").replace("_", " "))
            for k, v in fname_intel.items():
                if k not in meta:
                    meta[k] = v

    # Always mine the filename for intelligence
    name_text = p.stem.replace("-", " ").replace("_", " ")
    fname_intel = extract_intelligence(name_text)
    for k, v in fname_intel.items():
        if k not in meta or not meta[k]:
            meta[k] = v

    return meta

def scan_directory(root, max_files=2000, folders=None):
    results, errors = [], []
    count = 0
    scan_roots = [os.path.join(root, f) for f in folders] if folders else [root]

    for scan_root in scan_roots:
        if not os.path.exists(scan_root):
            print(f"  [SKIP] Not found: {scan_root}")
            continue
        for dirpath, dirs, files in os.walk(scan_root):
            dirs[:] = [d for d in dirs if not d.startswith(".") and d not in ["Android","data","obb"]]
            for fname in files:
                if fname.startswith("."):
                    continue
                fpath = os.path.join(dirpath, fname)
                try:
                    r = scan_file(fpath)
                    if r:
                        results.append(r)
                        count += 1
                        if count % 50 == 0:
                            print(f"  Scanned {count} files...")
                except Exception as e:
                    errors.append({"file": fpath, "error": str(e)})
                if count >= max_files:
                    return results, errors
    return results, errors

# ── Deduplication Index ───────────────────────────────────────────────────────
def load_master_index():
    if os.path.exists(MASTER_LOG):
        try:
            with open(MASTER_LOG) as f:
                data = json.load(f)
            fps = set()
            for entry in data.get("fingerprints", []):
                if isinstance(entry, str):
                    fps.add(entry)
                elif isinstance(entry, dict):
                    fp = f"{entry.get('file_path','')}|{entry.get('md5_hash','')}"
                    fps.add(fp)
            return fps, data.get("scan_history", []), data.get("records", [])
        except Exception:
            pass
    return set(), [], []

def save_master_index(fingerprints, scan_history, records):
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    with open(MASTER_LOG, "w") as f:
        json.dump({
            "fingerprints": list(fingerprints),
            "scan_history": scan_history,
            "records": records,
            "exported": datetime.now().isoformat(),
            "version": "2.0"
        }, f, indent=2)

def fingerprint(r):
    return f"{r.get('file_path','')}|{r.get('md5_hash','')}"

# ── Excel Export ──────────────────────────────────────────────────────────────
ALL_COLS = [
    "file_name","file_type","category","file_size","modified",
    "risk_flag","names_found","emails_found","phones_found",
    "addresses_found","orgs_found","case_numbers","dates_found",
    "keywords_hit","money_refs","ssns_found",
    "gps_lat","gps_lon","gps_location",
    "exif_Make","exif_Model","exif_DateTime","exif_DateTimeOriginal",
    "audio_artist","audio_title","audio_album","audio_duration","audio_year",
    "img_width","img_height",
    "content_preview","md5_hash","folder",
]
ALL_HDRS = [
    "File Name","Type","Category","Size (bytes)","Modified",
    "Risk Flag","Names Found","Emails","Phones",
    "Addresses","Orgs","Case Numbers","Dates","Keywords","Money Refs","SSNs",
    "GPS Lat","GPS Lon","GPS Link",
    "Camera Make","Camera Model","EXIF Date","EXIF Date Original",
    "Audio Artist","Audio Title","Album","Duration","Year",
    "Width","Height",
    "Content Preview","MD5","Folder",
]

def _hfill():
    return PatternFill(start_color="0F1628", fill_type="solid")
def _hfont():
    return Font(bold=True, color="00D4FF", size=10)

def _style_sheet(ws, widths=None):
    for cell in ws[1]:
        cell.fill  = _hfill()
        cell.font  = _hfont()
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    ws.freeze_panes = "A2"
    if widths:
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

RISK_FILLS = {
    "High":   PatternFill(start_color="C0392B", fill_type="solid"),
    "Medium": PatternFill(start_color="D68910", fill_type="solid"),
    "Low":    PatternFill(start_color="1E8449", fill_type="solid"),
}

def _clean(val):
    """Strip illegal XML/openpyxl characters from any cell value."""
    if not isinstance(val, str):
        return val
    import re as _re
    val = _re.sub(r'[\x00-\x08\x0B\x0C\x0E-\x1F\x7F]', '', val)
    val = val.replace('\x0D', ' ')
    return val[:32767]  # Excel cell max length

def _clean_row(row):
    return [_clean(v) for v in row]

def append_to_master(new_results, label, elapsed, existing_total):
    if not OPENPYXL_OK:
        print("[SKIP] openpyxl not available")
        return

    os.makedirs(OUTPUT_DIR, exist_ok=True)

    if os.path.exists(MASTER_XLSX):
        from openpyxl import load_workbook
        wb = load_workbook(MASTER_XLSX)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Summary"
        ws.append(["OSINT AI NEO — MASTER INTELLIGENCE SHEET"])
        ws["A1"].font = Font(bold=True, color="00D4FF", size=14)
        ws.append(["Last Updated:",""])
        ws.append(["Total Files:","0"])
        ws.append(["High Risk Files:","0"])
        ws.append(["Files with Names:","0"])
        ws.append(["Files with GPS:","0"])
        ws.append([])
        ws.append(["#","Folder Scanned","New Files","Date","Duration(s)"])
        for c in ws[8]: c.fill=_hfill(); c.font=_hfont()
        wb.create_sheet("All Files").append(ALL_HDRS)
        wb.create_sheet("High Risk").append(ALL_HDRS)
        wb.create_sheet("Names & Contacts").append(ALL_HDRS)
        wb.create_sheet("GPS Data").append(ALL_HDRS)
        wb.create_sheet("Emails & Phones").append(ALL_HDRS)
        for sh in ["All Files","High Risk","Names & Contacts","GPS Data","Emails & Phones"]:
            _style_sheet(wb[sh])

    ws_all   = wb["All Files"]
    ws_risk  = wb["High Risk"]
    ws_names = wb["Names & Contacts"]
    ws_gps   = wb["GPS Data"]
    ws_ep    = wb["Emails & Phones"]

    hi_count = names_count = gps_count = ep_count = 0

    for r in new_results:
        row = _clean_row([r.get(c, "") for c in ALL_COLS])
        ws_all.append(row)
        risk = r.get("risk_flag","")
        # Color risk cell (col 6 = F)
        last_row = ws_all.max_row
        if risk in RISK_FILLS:
            ws_all.cell(last_row, 6).fill = RISK_FILLS[risk]

        if risk == "High":
            ws_risk.append(row)
            hi_count += 1
        if r.get("names_found") or r.get("emails_found") or r.get("phones_found"):
            ws_names.append(row)
            names_count += 1
        if r.get("gps_lat"):
            ws_gps.append(row)
            gps_count += 1
        if r.get("emails_found") or r.get("phones_found"):
            ws_ep.append(row)
            ep_count += 1

    # Update Summary
    ws_sum = wb["Summary"]
    now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws_sum["B2"] = now_str
    ws_sum["B3"] = existing_total + len(new_results)
    ws_sum["B4"] = (ws_sum["B4"].value or 0) + hi_count
    ws_sum["B5"] = (ws_sum["B5"].value or 0) + names_count
    ws_sum["B6"] = (ws_sum["B6"].value or 0) + gps_count
    scan_num = ws_sum.max_row - 7
    ws_sum.append([f"#{scan_num+1}", label, len(new_results), now_str, elapsed])

    wb.save(MASTER_XLSX)
    print(f"  Sheets: All Files={ws_all.max_row-1} | High Risk={hi_count} | GPS={gps_count} | Names/Contacts={names_count}")

# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    print("\n" + "="*55)
    print("  OSINT AI NEO — Android Phone Scanner v2")
    print("  Deep Extraction: emails, phones, names, addresses,")
    print("  GPS, EXIF, audio tags, case numbers, risk keywords")
    print("  Read-only. Your files are never modified.")
    print("="*55 + "\n")

    fps, scan_history, old_records = load_master_index()
    if fps:
        print(f"  Master list: {len(fps)} files already recorded.")
        print(f"  Only NEW files will be added this scan.\n")
    else:
        print("  No master list yet — building from scratch.\n")

    print("Select scan target:")
    print("  1. Documents")
    print("  2. DCIM (Camera Photos + Videos)")
    print("  3. Downloads")
    print("  4. WhatsApp")
    print("  5. Contacts (VCF)")
    print("  6. ALL phone storage (full deep scan)")
    print("  7. Custom path")
    print()

    folder_map = {
        "1": ["Documents"], "2": ["DCIM"], "3": ["Download"],
        "4": ["WhatsApp"],  "5": ["Contacts"], "6": None, "7": "custom",
    }

    choice = input("Choice (1-7) [default 1]: ").strip() or "1"

    if choice == "7":
        custom = input(f"Path under {PHONE_ROOT}/: ").strip()
        scan_folders = [custom] if custom else ["Documents"]
    else:
        scan_folders = folder_map.get(choice, ["Documents"])

    max_str = input("Max files [default 2000]: ").strip()
    max_files = int(max_str) if max_str.isdigit() else 2000

    label = "All Storage" if scan_folders is None else ", ".join(scan_folders)
    print(f"\n  Scanning: {label}  |  Max: {max_files} files\n")

    start = datetime.now()
    results, errors = scan_directory(PHONE_ROOT, max_files=max_files, folders=scan_folders)
    elapsed = round((datetime.now() - start).total_seconds(), 1)

    new_results = [r for r in results if fingerprint(r) not in fps]
    dupe_count  = len(results) - len(new_results)

    print(f"\n  Scanned:   {len(results)} files in {elapsed}s")
    print(f"  New:       {len(new_results)} files")
    print(f"  Dupes:     {dupe_count} (already in master)")
    if errors:
        print(f"  Errors:    {len(errors)}")

    if not new_results:
        print("\n  ✓ Master list is fully up to date.\n")
        return

    # Intelligence summary
    hi = sum(1 for r in new_results if r.get("risk_flag") == "High")
    med = sum(1 for r in new_results if r.get("risk_flag") == "Medium")
    named = sum(1 for r in new_results if r.get("names_found"))
    emailed = sum(1 for r in new_results if r.get("emails_found"))
    phoned = sum(1 for r in new_results if r.get("phones_found"))
    gps_ct = sum(1 for r in new_results if r.get("gps_lat"))

    print(f"\n  Intelligence found in new files:")
    print(f"    🔴 High Risk:    {hi}")
    print(f"    🟡 Medium Risk:  {med}")
    print(f"    👤 Names:        {named} files")
    print(f"    📧 Emails:       {emailed} files")
    print(f"    📞 Phones:       {phoned} files")
    print(f"    📍 GPS:          {gps_ct} files")

    # Save index
    for r in new_results:
        fps.add(fingerprint(r))
    all_records = old_records + new_results
    scan_history.append({
        "label": label, "new_files": len(new_results),
        "date": datetime.now().isoformat(), "elapsed": elapsed,
        "high_risk": hi, "names": named, "gps": gps_ct,
    })
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    save_master_index(fps, scan_history, all_records)
    print(f"\n  ✓ Index:  {MASTER_LOG}  ({len(fps)} total files)")

    if OPENPYXL_OK:
        append_to_master(new_results, label, elapsed, len(fps) - len(new_results))
        print(f"  ✓ Excel: {MASTER_XLSX}")
    else:
        csv_path = f"{OUTPUT_DIR}/OSINT_Master.csv"
        write_hdr = not os.path.exists(csv_path)
        all_keys = list(dict.fromkeys(k for r in new_results for k in r.keys()))
        with open(csv_path, "a", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=all_keys, extrasaction="ignore")
            if write_hdr:
                writer.writeheader()
            writer.writerows(new_results)
        print(f"  ✓ CSV: {csv_path}")

    if errors:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        err_path = f"{OUTPUT_DIR}/errors_{ts}.json"
        with open(err_path, "w") as f:
            json.dump(errors, f, indent=2)

    print(f"\n{'='*55}")
    print(f"  Done! {len(new_results)} new files added.")
    print(f"  Total in master: {len(fps)} files")
    print(f"  Upload {MASTER_LOG} to OSINT AI Neo to sync.")
    print(f"{'='*55}\n")

if __name__ == "__main__":
    main()
