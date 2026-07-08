from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from datetime import datetime
import json

MASTER_PATH = "data/OSINT_Master_Sheet.xlsx"

HEADER_FILL = PatternFill(start_color="0F1628", fill_type="solid")
HEADER_FONT = Font(bold=True, color="00D4FF", size=11)
ACCENT_FILL = PatternFill(start_color="141D35", fill_type="solid")
RISK_HIGH = PatternFill(start_color="C0392B", fill_type="solid")
RISK_MED = PatternFill(start_color="D68910", fill_type="solid")
RISK_LOW = PatternFill(start_color="1E8449", fill_type="solid")
RISK_FONT = Font(bold=True, color="FFFFFF", size=9)
THIN_BORDER = Border(
    left=Side(style="thin", color="1E2D50"),
    right=Side(style="thin", color="1E2D50"),
    top=Side(style="thin", color="1E2D50"),
    bottom=Side(style="thin", color="1E2D50")
)

def style_header_row(ws, num_cols):
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER

def style_data_rows(ws, num_rows, num_cols, risk_col=None):
    for row_idx in range(2, num_rows + 2):
        for col_idx in range(1, num_cols + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if risk_col and col_idx == risk_col:
                if cell.value == "High":
                    cell.fill = RISK_HIGH
                    cell.font = RISK_FONT
                elif cell.value == "Medium":
                    cell.fill = RISK_MED
                    cell.font = RISK_FONT
                elif cell.value == "Low":
                    cell.fill = RISK_LOW
                    cell.font = RISK_FONT

def generate_master_sheet(entities, relationships, events, file_scans=None, scans=None):
    wb = Workbook()
    wb.remove(wb.active)

    # ── SUMMARY SHEET ──
    ws_sum = wb.create_sheet("📊 Summary")
    ws_sum.sheet_properties.tabColor = "00D4FF"
    ws_sum["A1"] = "OSINT AI NEO — MASTER INTELLIGENCE REPORT"
    ws_sum["A1"].font = Font(bold=True, color="00D4FF", size=14)
    ws_sum["A2"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws_sum["A2"].font = Font(color="5A7090", size=10)
    ws_sum.append([])
    ws_sum.append(["Category", "Count"])
    stats_data = [
        ("Entities", len(entities)),
        ("Relationships", len(relationships)),
        ("Events / Timeline", len(events)),
        ("File Scans", len(file_scans) if file_scans else 0),
        ("Target Scans", len(scans) if scans else 0),
        ("High Risk Entities", sum(1 for e in entities if e.get("risk_level") == "High")),
        ("Medium Risk Entities", sum(1 for e in entities if e.get("risk_level") == "Medium")),
        ("Low Risk Entities", sum(1 for e in entities if e.get("risk_level") == "Low")),
    ]
    for row in stats_data:
        ws_sum.append(row)
    for cell in ws_sum["A4:B4"][0]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
    for row_idx in range(5, 5 + len(stats_data)):
        for col_idx in range(1, 3):
            cell = ws_sum.cell(row=row_idx, column=col_idx)
            cell.border = THIN_BORDER
    ws_sum.column_dimensions["A"].width = 30
    ws_sum.column_dimensions["B"].width = 15

    # ── ENTITIES SHEET ──
    ws_ent = wb.create_sheet("🎯 Entities")
    headers = ["Entity ID", "Type", "Label", "Category", "Geo Location / Address", "Risk Level", "Source", "Notes", "Created At"]
    ws_ent.append(headers)
    for ent in entities:
        ws_ent.append([
            ent.get("entity_id", ""), ent.get("type", ""), ent.get("label", ""),
            ent.get("category", ""), ent.get("geo_location", ""), ent.get("risk_level", ""),
            ent.get("source", ""), ent.get("notes", ""), ent.get("created_at", "")
        ])
    style_header_row(ws_ent, len(headers))
    style_data_rows(ws_ent, len(entities), len(headers), risk_col=6)
    for i, w in enumerate([12, 12, 22, 18, 22, 12, 16, 40, 18], 1):
        ws_ent.column_dimensions[get_column_letter(i)].width = w
    ws_ent.freeze_panes = "A2"
    ws_ent.row_dimensions[1].height = 25

    # ── RELATIONSHIPS SHEET ──
    ws_rel = wb.create_sheet("🔗 Relationships")
    rel_headers = ["Relation ID", "Source Entity", "Target Entity", "Relationship Type", "Confidence", "Source", "Created At"]
    ws_rel.append(rel_headers)
    for rel in relationships:
        ws_rel.append([
            rel.get("relation_id", ""), rel.get("source_entity", ""), rel.get("target_entity", ""),
            rel.get("relationship_type", ""), rel.get("confidence", ""), rel.get("source", ""), rel.get("created_at", "")
        ])
    style_header_row(ws_rel, len(rel_headers))
    style_data_rows(ws_rel, len(relationships), len(rel_headers))
    for i, w in enumerate([12, 22, 28, 22, 12, 20, 18], 1):
        ws_rel.column_dimensions[get_column_letter(i)].width = w
    ws_rel.freeze_panes = "A2"

    # ── EVENTS SHEET ──
    ws_ev = wb.create_sheet("📅 Events")
    ev_headers = ["Event ID", "Date/Timestamp", "Event Type", "Location", "Entities Involved", "Source", "Created At"]
    ws_ev.append(ev_headers)
    for ev in events:
        ws_ev.append([
            ev.get("event_id", ""), ev.get("timestamp", ""), ev.get("event_type", ""),
            ev.get("location", ""), ev.get("entities_involved", ""), ev.get("source", ""), ev.get("created_at", "")
        ])
    style_header_row(ws_ev, len(ev_headers))
    style_data_rows(ws_ev, len(events), len(ev_headers))
    for i, w in enumerate([12, 18, 22, 18, 22, 22, 18], 1):
        ws_ev.column_dimensions[get_column_letter(i)].width = w
    ws_ev.freeze_panes = "A2"

    # ── FILE SCANS SHEET ──
    if file_scans:
        ws_fs = wb.create_sheet("📁 File Scans")
        fs_headers = ["File Path", "File Type", "File Size (bytes)", "Key Metadata", "Scanned At"]
        ws_fs.append(fs_headers)
        for fs in file_scans:
            try:
                meta = json.loads(fs.get("metadata_json", "{}"))
                meta_str = " | ".join(f"{k}: {v}" for k, v in list(meta.items())[:6])
            except Exception:
                meta_str = ""
            ws_fs.append([
                fs.get("file_path", ""), fs.get("file_type", ""), fs.get("file_size", 0),
                meta_str, fs.get("created_at", "")
            ])
        style_header_row(ws_fs, len(fs_headers))
        style_data_rows(ws_fs, len(file_scans), len(fs_headers))
        for i, w in enumerate([45, 12, 18, 50, 18], 1):
            ws_fs.column_dimensions[get_column_letter(i)].width = w
        ws_fs.freeze_panes = "A2"

    # ── TARGET SCANS SHEET ──
    if scans:
        ws_sc = wb.create_sheet("🔍 Target Scans")
        sc_headers = ["Target", "Scan Type", "Results Summary", "Scanned At"]
        ws_sc.append(sc_headers)
        for sc in scans:
            try:
                result = json.loads(sc.get("result_json", "{}"))
                result_str = " | ".join(f"{k}: {v}" for k, v in list(result.items())[:5])
            except Exception:
                result_str = ""
            ws_sc.append([sc.get("target", ""), sc.get("scan_type", ""), result_str, sc.get("created_at", "")])
        style_header_row(ws_sc, len(sc_headers))
        style_data_rows(ws_sc, len(scans), len(sc_headers))
        for i, w in enumerate([25, 15, 60, 18], 1):
            ws_sc.column_dimensions[get_column_letter(i)].width = w
        ws_sc.freeze_panes = "A2"

    import os
    os.makedirs("data", exist_ok=True)
    wb.save(MASTER_PATH)
    return MASTER_PATH
