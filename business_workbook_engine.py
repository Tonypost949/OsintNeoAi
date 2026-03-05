#!/usr/bin/env python3
"""
Business Workbook Engine
========================
Builds a complete multi-sheet Excel workbook with:
  - Pre-built headers on every sheet
  - Autofilled example rows
  - Data validation dropdowns (Entity Type, Status, License Status)
  - Formatted headers, column widths, and filters
  - Reusable: run anytime to regenerate a clean workbook

Usage:
    python3 business_workbook_engine.py              # build default workbook
    python3 business_workbook_engine.py -o my.xlsx   # custom output name
"""

import argparse
import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ── Palette ──────────────────────────────────────────────────────────────────
HEADER_FONT = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
ALT_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)

# ── Validation option lists ──────────────────────────────────────────────────
ENTITY_TYPES = [
    "Sole Proprietorship", "LLC", "Corporation", "Partnership",
    "Nonprofit", "Trust", "Government", "Joint Venture", "Other",
]
STATUSES = [
    "Active", "Inactive", "Pending", "Suspended", "Dissolved", "Under Review",
]
LICENSE_STATUSES = [
    "Valid", "Expired", "Pending Renewal", "Revoked", "Suspended", "Not Required",
]
CONTACT_ROLES = [
    "Owner", "Officer", "Registered Agent", "Director",
    "Accountant", "Attorney", "Compliance", "Other",
]
COMPLIANCE_TYPES = [
    "Annual Report", "Tax Filing", "License Renewal",
    "Audit", "Inspection", "Registration Update", "Other",
]
COMPLIANCE_RESULT = ["Pass", "Fail", "Pending", "N/A"]

# ── Sheet definitions ────────────────────────────────────────────────────────
SHEETS = {
    "Entities": {
        "headers": [
            "Entity ID", "Legal Name", "DBA / Trade Name", "Entity Type",
            "FEIN / TIN", "State of Formation", "Formation Date",
            "Status", "Address", "City", "State", "ZIP",
            "Phone", "Email", "Website", "Notes",
        ],
        "widths": [12, 28, 24, 20, 18, 18, 14, 14, 30, 16, 8, 10, 16, 26, 26, 30],
        "rows": [
            ["E-001", "Acme Holdings LLC", "Acme Co", "LLC", "12-3456789",
             "Delaware", "2019-03-15", "Active", "100 Main St", "Wilmington",
             "DE", "19801", "(302) 555-0100", "info@acme.example.com",
             "https://acme.example.com", "Primary operating entity"],
            ["E-002", "Bright Future Nonprofit Inc", "", "Nonprofit", "98-7654321",
             "California", "2021-07-01", "Active", "200 Oak Ave", "Los Angeles",
             "CA", "90001", "(213) 555-0200", "hello@brightfuture.example.org",
             "https://brightfuture.example.org", "501(c)(3) approved"],
            ["E-003", "Summit Partners LP", "Summit LP", "Partnership", "55-1234567",
             "New York", "2017-11-20", "Under Review", "300 Wall St", "New York",
             "NY", "10005", "(212) 555-0300", "ops@summit.example.com",
             "", "Annual report overdue"],
        ],
    },
    "Contacts": {
        "headers": [
            "Contact ID", "Entity ID", "Full Name", "Role",
            "Title", "Phone", "Email", "LinkedIn",
            "Primary Contact", "Notes",
        ],
        "widths": [12, 12, 24, 18, 20, 16, 28, 30, 16, 30],
        "rows": [
            ["C-001", "E-001", "Jane Doe", "Owner", "CEO",
             "(302) 555-0101", "jane@acme.example.com",
             "linkedin.com/in/janedoe", "Yes", "Founding member"],
            ["C-002", "E-001", "John Smith", "Registered Agent", "General Counsel",
             "(302) 555-0102", "john@acme.example.com",
             "", "No", "Also handles compliance"],
            ["C-003", "E-002", "Maria Garcia", "Officer", "Executive Director",
             "(213) 555-0201", "maria@brightfuture.example.org",
             "linkedin.com/in/mariagarcia", "Yes", "Board liaison"],
            ["C-004", "E-003", "David Lee", "Owner", "Managing Partner",
             "(212) 555-0301", "david@summit.example.com",
             "linkedin.com/in/davidlee", "Yes", ""],
        ],
    },
    "Licenses": {
        "headers": [
            "License ID", "Entity ID", "License Type", "Issuing Authority",
            "License Number", "Issue Date", "Expiration Date",
            "License Status", "Renewal Notes",
        ],
        "widths": [12, 12, 24, 26, 20, 14, 14, 16, 30],
        "rows": [
            ["L-001", "E-001", "Business License", "City of Wilmington",
             "BL-2024-44012", "2024-01-10", "2025-01-10",
             "Valid", "Auto-renew enabled"],
            ["L-002", "E-001", "Sales Tax Permit", "Delaware Dept of Revenue",
             "ST-887766", "2019-03-20", "N/A",
             "Valid", "No expiration"],
            ["L-003", "E-002", "Charitable Solicitation", "CA Attorney General",
             "CS-2021-5500", "2021-08-01", "2025-08-01",
             "Pending Renewal", "Renewal submitted 2025-06-15"],
            ["L-004", "E-003", "Investment Adviser", "SEC",
             "IA-334455", "2017-12-01", "2024-12-01",
             "Expired", "Must renew before operating"],
        ],
    },
    "Compliance": {
        "headers": [
            "Record ID", "Entity ID", "Compliance Type", "Due Date",
            "Completed Date", "Result", "Responsible Contact",
            "Filing Reference", "Notes",
        ],
        "widths": [12, 12, 20, 14, 14, 12, 22, 20, 30],
        "rows": [
            ["CR-001", "E-001", "Annual Report", "2025-03-01", "2025-02-20",
             "Pass", "Jane Doe", "AR-2025-001", "Filed on time"],
            ["CR-002", "E-002", "Tax Filing", "2025-04-15", "",
             "Pending", "Maria Garcia", "", "990 due"],
            ["CR-003", "E-003", "Annual Report", "2024-11-30", "",
             "Fail", "David Lee", "", "Missed deadline — suspension risk"],
        ],
    },
    "Activity Log": {
        "headers": [
            "Log ID", "Entity ID", "Date", "Action",
            "Performed By", "Details",
        ],
        "widths": [12, 12, 14, 28, 20, 40],
        "rows": [
            ["LOG-001", "E-001", "2025-02-20", "Filed annual report",
             "Jane Doe", "Submitted via SOS online portal"],
            ["LOG-002", "E-002", "2025-06-15", "License renewal submitted",
             "Maria Garcia", "Charitable solicitation renewal for CA"],
            ["LOG-003", "E-003", "2025-01-05", "Status flagged for review",
             "System", "Annual report overdue > 30 days"],
        ],
    },
    "Validation Lists": {
        "headers": [
            "Entity Types", "Statuses", "License Statuses",
            "Contact Roles", "Compliance Types", "Compliance Results",
        ],
        "widths": [22, 18, 20, 18, 22, 20],
        "rows": [],  # filled programmatically below
    },
}

# Build the Validation Lists rows from the option lists
_max_len = max(
    len(ENTITY_TYPES), len(STATUSES), len(LICENSE_STATUSES),
    len(CONTACT_ROLES), len(COMPLIANCE_TYPES), len(COMPLIANCE_RESULT),
)
for i in range(_max_len):
    SHEETS["Validation Lists"]["rows"].append([
        ENTITY_TYPES[i] if i < len(ENTITY_TYPES) else "",
        STATUSES[i] if i < len(STATUSES) else "",
        LICENSE_STATUSES[i] if i < len(LICENSE_STATUSES) else "",
        CONTACT_ROLES[i] if i < len(CONTACT_ROLES) else "",
        COMPLIANCE_TYPES[i] if i < len(COMPLIANCE_TYPES) else "",
        COMPLIANCE_RESULT[i] if i < len(COMPLIANCE_RESULT) else "",
    ])


# ── Builder ──────────────────────────────────────────────────────────────────
def _style_header(ws, col_count):
    """Apply formatting to header row."""
    for col in range(1, col_count + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


def _set_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _add_autofilter(ws, col_count):
    ws.auto_filter.ref = f"A1:{get_column_letter(col_count)}1"


def _stripe_rows(ws, row_count, col_count):
    """Alternate row shading for readability."""
    for r in range(2, row_count + 2):
        for c in range(1, col_count + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if r % 2 == 0:
                cell.fill = ALT_FILL


def _add_validation(ws, col_idx, formula, row_start=2, row_end=200):
    """Add a dropdown data-validation rule to a column range."""
    col_letter = get_column_letter(col_idx)
    dv = DataValidation(type="list", formula1=formula, allow_blank=True)
    dv.error = "Please select a value from the dropdown list."
    dv.errorTitle = "Invalid Entry"
    dv.prompt = "Choose from the list."
    dv.promptTitle = "Selection"
    dv.sqref = f"{col_letter}{row_start}:{col_letter}{row_end}"
    ws.add_data_validation(dv)


def build_workbook(output_path: str) -> str:
    """Build the complete business workbook and return the file path."""
    wb = Workbook()
    first = True

    for sheet_name, spec in SHEETS.items():
        if first:
            ws = wb.active
            ws.title = sheet_name
            first = False
        else:
            ws = wb.create_sheet(title=sheet_name)

        headers = spec["headers"]
        ws.append(headers)
        for row in spec["rows"]:
            ws.append(row)

        col_count = len(headers)
        _style_header(ws, col_count)
        _set_widths(ws, spec["widths"])
        _add_autofilter(ws, col_count)
        _stripe_rows(ws, len(spec["rows"]), col_count)

    # Freeze top row on every sheet
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"

    # ── Data validations (reference the Validation Lists sheet) ──────────
    vl_title = "'Validation Lists'"
    entity_count = len(ENTITY_TYPES)
    status_count = len(STATUSES)
    lic_count = len(LICENSE_STATUSES)
    role_count = len(CONTACT_ROLES)
    comp_type_count = len(COMPLIANCE_TYPES)
    comp_res_count = len(COMPLIANCE_RESULT)

    # Entities sheet
    ent_ws = wb["Entities"]
    _add_validation(ent_ws, 4, f"={vl_title}!$A$2:$A${entity_count + 1}")   # Entity Type
    _add_validation(ent_ws, 8, f"={vl_title}!$B$2:$B${status_count + 1}")    # Status

    # Contacts sheet
    con_ws = wb["Contacts"]
    _add_validation(con_ws, 4, f"={vl_title}!$D$2:$D${role_count + 1}")      # Role
    _add_validation(con_ws, 9, '"Yes,No"')                                     # Primary Contact

    # Licenses sheet
    lic_ws = wb["Licenses"]
    _add_validation(lic_ws, 8, f"={vl_title}!$C$2:$C${lic_count + 1}")       # License Status

    # Compliance sheet
    cmp_ws = wb["Compliance"]
    _add_validation(cmp_ws, 3, f"={vl_title}!$E$2:$E${comp_type_count + 1}") # Compliance Type
    _add_validation(cmp_ws, 6, f"={vl_title}!$F$2:$F${comp_res_count + 1}")  # Result

    # ── Summary sheet (quick-reference counts) ───────────────────────────
    summary = wb.create_sheet(title="Summary", index=0)
    summary.append(["Business Workbook Summary"])
    summary.merge_cells("A1:C1")
    summary["A1"].font = Font(name="Calibri", bold=True, size=14, color="2F5496")
    summary["A1"].alignment = Alignment(horizontal="center")

    summary.append([])
    summary.append(["Sheet", "Records", "Description"])
    for c in range(1, 4):
        cell = summary.cell(row=3, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER

    info_rows = [
        ("Entities", len(SHEETS["Entities"]["rows"]), "Business entities & organizations"),
        ("Contacts", len(SHEETS["Contacts"]["rows"]), "People linked to entities"),
        ("Licenses", len(SHEETS["Licenses"]["rows"]), "Permits, licenses & registrations"),
        ("Compliance", len(SHEETS["Compliance"]["rows"]), "Filing & compliance tracking"),
        ("Activity Log", len(SHEETS["Activity Log"]["rows"]), "Chronological action history"),
        ("Validation Lists", _max_len, "Dropdown source values"),
    ]
    for i, (name, count, desc) in enumerate(info_rows, start=4):
        summary.append([name, count, desc])
        for c in range(1, 4):
            summary.cell(row=i, column=c).border = THIN_BORDER
        if i % 2 == 0:
            for c in range(1, 4):
                summary.cell(row=i, column=c).fill = ALT_FILL

    summary.append([])
    summary.append(["Generated", datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    summary.column_dimensions["A"].width = 22
    summary.column_dimensions["B"].width = 12
    summary.column_dimensions["C"].width = 40

    wb.save(output_path)
    return output_path


# ── CLI ──────────────────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(
        description="Build a complete multi-sheet business workbook."
    )
    parser.add_argument(
        "-o", "--output",
        default="business_workbook.xlsx",
        help="Output file path (default: business_workbook.xlsx)",
    )
    args = parser.parse_args()
    path = build_workbook(args.output)
    print(f"✅ Workbook created: {path}")
    print(f"   Sheets: {', '.join(SHEETS.keys())}, Summary")
    print("   Validation dropdowns: Entity Type, Status, License Status, Role, Compliance Type, Result")
    print("   Run again anytime to regenerate a fresh copy.")


if __name__ == "__main__":
    main()
