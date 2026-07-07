#!/usr/bin/env python3
"""
OSINT Suite Utilities
=====================
Shared components, base classes, and helper functions for the OSINT suite.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ── Excel Palette ────────────────────────────────────────────────────────────
HEADER_FONT = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
ALT_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)


class BaseWorkbookGenerator:
    """Base class for generating styled Excel workbooks."""

    def __init__(self):
        self.wb = Workbook()

    def _style_header(self, ws, col_count, fill=HEADER_FILL, start_row=1):
        for col in range(1, col_count + 1):
            cell = ws.cell(row=start_row, column=col)
            cell.font = HEADER_FONT
            cell.fill = fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = THIN_BORDER

    def _set_widths(self, ws, widths):
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

    def _add_autofilter(self, ws, col_count, row_count=1):
        ws.auto_filter.ref = f"A1:{get_column_letter(col_count)}{row_count}"

    def _stripe_rows(self, ws, row_count, col_count):
        for r in range(2, row_count + 2):
            for c in range(1, col_count + 1):
                cell = ws.cell(row=r, column=c)
                cell.border = THIN_BORDER
                cell.alignment = Alignment(vertical="center", wrap_text=True)
                if r % 2 == 0:
                    cell.fill = ALT_FILL

    def _add_validation(self, ws, col_idx, formula, row_start=2, row_end=500):
        col_letter = get_column_letter(col_idx)
        dv = DataValidation(type="list", formula1=formula, allow_blank=True)
        dv.error = "Please select a value from the dropdown list."
        dv.errorTitle = "Invalid Entry"
        dv.sqref = f"{col_letter}{row_start}:{col_letter}{row_end}"
        ws.add_data_validation(dv)
